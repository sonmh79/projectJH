import sys
import os
import datetime
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import uic
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager

#UI파일 연결
#단, UI파일은 Python 코드 파일과 같은 디렉토리에 위치해야한다.
form_class = uic.loadUiType("jh.ui")[0]
## python실행파일 디렉토리
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
form_class = uic.loadUiType(BASE_DIR + r'/jh.ui')[0]

#화면을 띄우는데 사용되는 Class 선언
class WindowClass(QMainWindow, form_class) :
    def __init__(self) :
        super().__init__()
        self.setupUi(self)
        self.driver = webdriver.Chrome(ChromeDriverManager().install())

        # Fixed Variable
        # 박주환 바보
        self.AE10 = {"MOGENS MAERSK": "1RM", "MARSEILLE MAERSK":"Y29", "MARIE MAERSK" : "1JM", "MAJESTIC MAERSK":"1HM", "MADISON MAERSK":"1KM","MATHILDE MAERSK":"2BM", "MANCHESTER MAERSK":"Y30","MARY MAERSK":"1IM","MARIT MAERSK":"2AM","MAGLEBY MAERSK":"1LM","MAYVIEW MAERSK":"1PM"}
        self.AE05 = {"MAASTRICHT MAERSK":"Y34","MURCIA MAERSK":"Y31","METTE MAERSK":"1ZM","MUNICH MAERSK":"Y25","MERETE MAERSK":"1QM","MADRID MAERSK":"Y24","MARGRETHE MAERSK":"1XM","MILAN MAERSK":"Y27","MSC RIFAYA":"F6M"}
        self.vessel_codes = dict(self.AE10, **self.AE05)
        self.target_ports = ["Gdansk", "Bremerhaven"]
        self.months = {"Jan":1, "Feb":2, "Mar":3, "Apr":4, "May":5, "Jun":6, "Jul":7, "Aug":8, "Sep":9, "Oct":10, "Nov":11, "Dec":12 }
        self.filename = "test.xlsx"
        self.wb = openpyxl.load_workbook(filename=self.filename, read_only=False, data_only=True)
        self.ws = self.wb[self.wb.sheetnames[0]]

        # Vessel Variables ( Updated from Clicked Cell )
        self.selected_vessel = "MARIE MAERSK"
        self.vessel_w = ""

        # Date Variables
        self.date = QDate.currentDate()
        self.cur_date = self.date.toString("yyyy-MM-dd")
        self.dateEdit.setDate(self.date)
        self.timerVar = QTimer()
        self.interval = 60000
        self.cnt = 0

        # Others
        self.is_AE10 = False
        self.is_AE05 = False

        self.url = f"http://www.maersk.com/schedules/vesselSchedules?vesselCode={self.vessel_codes[self.selected_vessel]}&fromDate={self.cur_date}"
        self.driver.get(self.url)
        self.driver.implicitly_wait(10) # Wait for Pop-up Screen

        try:
            btn_cookies = self.driver.find_element(By.CLASS_NAME, "coi-banner__accept")
            btn_cookies.click()
        except:
            pass

        self.initTable()

        # Connect Functions with Widgets
        self.btn_update.clicked.connect(self.crawl)
        self.dateEdit.dateChanged.connect(self.dateChanged)
        self.btn_refresh.clicked.connect(self.initTable)
        self.btn_updateValue.clicked.connect(self.updateValue)
        # self.btn_start.clicked.connect(self.setTimer)
        # self.btn_stop.clicked.connect(self.stopTimer)
        # self.cmb_interval.currentIndexChanged.connect(self.changeInterval)

    def updateValue(self):

        """Update Value Modified by User"""

        excel_row = self.r + 2
        excel_col = self.c + 1
        target = get_column_letter(excel_col) + str(excel_row)
        newValue = self.valueEdit.text()
        self.ws[target] = newValue
        self.wb.save(self.filename)
        self.initTable()

    def resetUpdateLayout(self):

        """ Clear All Widgets in Update Layout """

        for i in reversed(range(self.updateLayout.count())):
            self.updateLayout.itemAt(i).widget().setParent(None)

    def showClickedLabel(self, r, c, t1 = True):

        """Show Clicked Label"""

        self.resetUpdateLayout()

        clickedColumnValue = self.df.iloc[r,c]
        vessel_info = self.df.iloc[r,1].split()
        self.selected_vessel = vessel_info[0] + " " + vessel_info[1]
        self.vessel_w = vessel_info[-1].strip(" ") # 142W
        self.lbl_vesselName.setText(f"Vessel Name : {self.selected_vessel}")
        self.lbl_vesselW.setText(f"Vessel W : {self.vessel_w}")

        layout = self.updateLayout

        e = QLineEdit()
        e.setText(str(clickedColumnValue))
        layout.addWidget(QLabel("Values: "))
        layout.addWidget(e)

        self.r, self.c, self.valueEdit = r, c, e


    def setLabel1(self, row, column):

        """Change Table1 Index to Dataframe Index"""

        r,c = row + self.visible_rows[0][0]-2, column + 1
        self.showClickedLabel(r,c,t1=True)

        self.is_AE10 = True
        self.is_AE05 = False

    def setLabel2(self, row, column):

        """Change Table2 Index to Dataframe Index"""

        r, c = row + self.visible_rows[1][0]-2, column + 1
        self.showClickedLabel(r, c, t1 = False)

        self.is_AE10 = False
        self.is_AE05 = True

    def initTable(self):

        """Initialize Table"""

        df = pd.read_excel(self.filename)
        ws = self.ws
        visible_rows = [[],[]]
        flag = True
        for i in range(4,len( df)+1):
            if tuple(ws.row_dimensions[i])[0][0] != "hidden":
                cell_value = ws.cell(row=i,column=2).value
                if cell_value != None:
                    if "MAERSK" in cell_value or "Blank" in cell_value:
                        if flag:
                            visible_rows[0].append(i)
                        else:
                            visible_rows[1].append(i)
                    else:
                        flag = False

        self.visible_rows = visible_rows

        table1 = self.table1
        table2 = self.table2

        df1 = df.iloc[visible_rows[0][0]-2:visible_rows[0][-1]+1-2,1:]
        df2 = df.iloc[visible_rows[1][0]-2:,1:8]
        df1.columns = ["Vessel Name(AE10)","Planned_ETA(B)","Current_ETA1(B)","Current_ETA2(B)","Delay(B)","Planned_ETA(G)","Current_ETA(G)","Delay(G)","Vessel Location","ETA Change"]
        df2.columns = ["Vessel Name(AE05)","Planned_ETA(B)","Current_ETA1(B)","Current_ETA2(B)","Delay(B)","Vessel Location","ETA Change"]
        self.df,self.df1,self.df2 = df,df1,df2

        table1.setColumnCount(len(df1.columns))
        table1.setRowCount(len(df1))
        for i in range(len(df1)):
            for j in range(len(df1.columns)):
                table1.setItem(i, j, QTableWidgetItem(str(df1.iloc[i, j])))
        table1.setHorizontalHeaderLabels(df1.columns)  # 컬럼 헤더 입력
        table1.resizeColumnsToContents()
        table1.cellClicked.connect(self.setLabel1)

        table2.setColumnCount(len(df2.columns))
        table2.setRowCount(len(df2))
        for i in range(len(df2)):
            for j in range(len(df2.columns)):
                table2.setItem(i, j, QTableWidgetItem(str(df2.iloc[i, j])))
        table2.setHorizontalHeaderLabels(df2.columns)  # 컬럼 헤더 입력
        table2.resizeColumnsToContents()
        table2.cellClicked.connect(self.setLabel2)


    # def changeInterval(self):
    #
    #     """Set Auto-Crawling Interval"""
    #
    #     text = self.cmb_interval.currentText()
    #     if text == "1분":
    #         self.interval = 1000 * 60
    #     elif text == "10분":
    #         self.interval = 1000 * 60 * 10
    #     elif text == "30분":
    #         self.interval = 1000 * 60 * 30
    #     elif text == "1시간":
    #         self.interval = 1000 * 60 * 60
    #     elif text == "10시간":
    #         self.interval = 1000 * 60 * 60 * 10
    #
    # def stopTimer(self):
    #
    #     """Stop Timer"""
    #
    #     self.timerVar.stop()
    #     self.txt_status.setText("OFF")
    #
    # def setTimer(self):
    #
    #     """Set Timer"""
    #
    #     self.timerVar.setInterval(self.interval)
    #     self.timerVar.timeout.connect(self.crawl)
    #     self.timerVar.start()
    #     self.txt_status.setText("ON")

    def dateChanged(self):

        """Set Changed Date"""

        self.date = self.dateEdit.date()
        self.cur_date = self.date.toString("yyyy-MM-dd")

    def message_question(self,title,text):

        """ Qmessagebox - Question """

        reply = QMessageBox.question(self, title, text,
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        return reply

    def message_warning(self,title,text):

        """ Qmessagebox - Question """

        reply = QMessageBox.warning(self, title, text,
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        return reply

    def crawl(self):
                                                                                1111
        """Start Crawling"""

        if self.selected_vessel not in self.vessel_codes:
            return self.message_warning("","Please Click the Vessel First")

        qdate = self.date
        driver = self.driver

        flag = True
        search_BGB = False
        BGB_result = []
        crawl_cnt = 0
        while flag:
            crawl_cnt += 1

            if crawl_cnt >= 20:
                return self.message_warning("","Cannot Find Data..")

            date = qdate.toString("yyyy-MM-dd")
            self.url = f"http://www.maersk.com/schedules/vesselSchedules?vesselCode={self.vessel_codes[self.selected_vessel]}&fromDate={date}"
            driver.get(self.url)
            # Get Divs from Schedule Result
            results = driver.find_elements(By.CLASS_NAME, "ptp-results__transport-plan--item")
            final_element = driver.find_elements(By.CLASS_NAME, "ptp-results__transport-plan--item-final")

            # Get Port Information like port name, a_or_d info, planned date

            for i, result in enumerate(results + final_element):
                port, _ = result.find_element(By.CLASS_NAME, "location").find_elements(By.TAG_NAME, "div")
                a_or_d, date = result.find_element(By.CLASS_NAME, "transport-label").find_elements(By.TAG_NAME, "div")

                p, d = port.text, date.text
                a_or_d_list = a_or_d.text.split("-")
                cur_vessel_w = a_or_d_list[-1].strip(" ") # ex) 143W

                if cur_vessel_w == self.vessel_w and p in ["Algeciras", "Suez Canal"]:
                    flag = False
                    search_BGB = True

                if search_BGB and p in self.target_ports:
                    port_name = p
                    date = d
                    day, month, year, time = date.split()
                    str_date = f"{year}-{self.months[month]}-{day} {time}:00"
                    trans_datetime = datetime.datetime.strptime(str_date,'%Y-%m-%d %H:%M:%S')
                    BGB_result.append((port_name, trans_datetime))

            # Page Search Interval - 7 days
            qdate = qdate.addDays(-7)

        if len(BGB_result) == 2:
            if BGB_result[0][0] == "Gdansk":
                BGB_result = [("Bremerhaven", "Omit")] + BGB_result
            else:
                BGB_result += [("Bremerhaven", "Not Yet")]
        text = self.selected_vessel + self.vessel_w + "\n"

        for result in BGB_result:
            text += f"{result[0]} - {result[1]} \n"

        reply = self.message_question("", text)
        if reply == QMessageBox.Yes:
            ws = self.ws
            excel_row = self.r + 2
            if self.is_AE10 and not self.is_AE05:
                brem_1 = "D" + str(excel_row)
                gdansk = "H" + str(excel_row)
                brem_2 = "E" + str(excel_row)

                ws[brem_1] = BGB_result[0][1]
                ws[gdansk] = BGB_result[1][1]
                ws[brem_2] = BGB_result[2][1]

            elif not self.is_AE10 and self.is_AE05:
                brem_1 = "D" + str(excel_row)
                brem_2 = "E" + str(excel_row)

                ws[brem_1] = BGB_result[0][1]
                ws[brem_2] = BGB_result[1][1]

            delay_cell = "F" + str(excel_row)
            try:
                delay = self.ws["D"+str(excel_row)].value - self.ws["C"+str(excel_row)].value
            except:
                delay = self.ws["E"+str(excel_row)].value - self.ws["C"+str(excel_row)].value

            day = delay.days
            if delay.seconds / (60 * 60 * 24) >= 0.5: day += 1
            ws[delay_cell] = day

            self.wb.save(self.filename)
            self.initTable()
        else:
            print("Not recorded")




if __name__ == "__main__" :
    app = QApplication(sys.argv)
    myWindow = WindowClass()
    myWindow.show()
    app.exec_()
    myWindow.driver.quit()